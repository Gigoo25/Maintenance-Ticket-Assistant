from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pynput.keyboard import Key, Controller
from pynput.keyboard import Key, Listener
from pathlib import Path
from PIL import Image
import openpyxl
from openpyxl import Workbook
import PySimpleGUI as sg
import csv
import time
import os
import sys

###########################################################

#get script directory
dir_path = os.path.dirname(os.path.abspath(__file__))
#set controller variable
keyboard = Controller()

###########################################################

def quit_chrome():
	browser.quit()
	time.sleep(1)
	browserExe = "Chrome"

def quit_chrome_exception():
	#### This exception occurs if the element are not found in the webpage.
	print("###########################################################")
	print("Error: An element was not found.")
	print("###########################################################")
	print("Exiting chrome in 10 seconds.")
	print("###########################################################")
	print('\n')
	#### Quit Browser
	time.sleep(10)
	quit_chrome()

# Function to convert   
def listToString(s):  
    
    # initialize an empty string 
    str1 = " " 
    
    # return string   
    return (str1.join(s)) 

###########################################################

# Delete old files if they exist
ModemReportCSV = dir_path+"\Maintenance Ticket Files\ModemReport.csv"
ModemReportCSV1 = dir_path+"\Maintenance Ticket Files\ModemReport (1).csv"
ModemReport_Red = dir_path+"\Maintenance Ticket Files\ModemReport_Red.xlsx"
ModemReport_RedYellow = dir_path+"\Maintenance Ticket Files\ModemReport_RedYellow.xlsx"
Full_Map_Element = dir_path+"\Maintenance Ticket Files\Full_Map_Element.png"
PEA_critical_modem_map = dir_path+"\Maintenance Ticket Files\PEA critical modem map.png"
PEA_critical_modems_XLSX = dir_path+"\Maintenance Ticket Files\PEA critical modems.xlsx"
PEA_vTDR_XLSX = dir_path+"\Maintenance Ticket Files\PEA vTDR.xlsx"

if os.path.isfile(ModemReportCSV) and os.access(ModemReportCSV, os.R_OK):
	os.remove(ModemReportCSV)

if os.path.isfile(ModemReportCSV1) and os.access(ModemReportCSV1, os.R_OK):
	os.remove(ModemReportCSV1)

if os.path.isfile(ModemReport_Red) and os.access(ModemReport_Red, os.R_OK):
	os.remove(ModemReport_Red)

if os.path.isfile(ModemReport_RedYellow) and os.access(ModemReport_RedYellow, os.R_OK):
	os.remove(ModemReport_RedYellow)

if os.path.isfile(Full_Map_Element) and os.access(Full_Map_Element, os.R_OK):
	os.remove(Full_Map_Element)

if os.path.isfile(PEA_critical_modem_map) and os.access(PEA_critical_modem_map, os.R_OK):
	os.remove(PEA_critical_modem_map)

if os.path.isfile(PEA_critical_modems_XLSX) and os.access(PEA_critical_modems_XLSX, os.R_OK):
	os.remove(PEA_critical_modems_XLSX)

if os.path.isfile(PEA_vTDR_XLSX) and os.access(PEA_vTDR_XLSX, os.R_OK):
	os.remove(PEA_vTDR_XLSX)

Full_Spectrum_MAX_Drop = dir_path+"\Maintenance Ticket Files\Full_Spectrum_Max_Element.png"
Full_Waterfall_MAX_Drop = dir_path+"\Maintenance Ticket Files\Full_Waterfall_Max_Element.png"

Full_Spectrum_AVG_Drop = dir_path+"\Maintenance Ticket Files\Full_Spectrum_Average_Element.png"
Full_Waterfall_AVG_Drop = dir_path+"\Maintenance Ticket Files\Full_Waterfall_Avg_Element.png"

Spectrum_MAX_Drop = dir_path+"\Maintenance Ticket Files\Spectrum_Max.png"
Waterfall_MAX_Drop = dir_path+"\Maintenance Ticket Files\Waterfall_Max.png"

Spectrum_AVG_Drop = dir_path+"\Maintenance Ticket Files\Spectrum_Average.png"
Waterfall_AVG_Drop = dir_path+"\Maintenance Ticket Files\Waterfall_Average.png"

if os.path.isfile(Full_Spectrum_MAX_Drop) and os.access(Full_Spectrum_MAX_Drop, os.R_OK):
	os.remove(Full_Spectrum_MAX_Drop)

if os.path.isfile(Full_Waterfall_MAX_Drop) and os.access(Full_Waterfall_MAX_Drop, os.R_OK):
	os.remove(Full_Waterfall_MAX_Drop)

if os.path.isfile(Full_Spectrum_AVG_Drop) and os.access(Full_Spectrum_AVG_Drop, os.R_OK):
	os.remove(Full_Spectrum_AVG_Drop)

if os.path.isfile(Full_Waterfall_AVG_Drop) and os.access(Full_Waterfall_AVG_Drop, os.R_OK):
	os.remove(Full_Waterfall_AVG_Drop)

if os.path.isfile(Spectrum_MAX_Drop) and os.access(Spectrum_MAX_Drop, os.R_OK):
	os.remove(Spectrum_MAX_Drop)

if os.path.isfile(Waterfall_MAX_Drop) and os.access(Waterfall_MAX_Drop, os.R_OK):
	os.remove(Waterfall_MAX_Drop)

if os.path.isfile(Spectrum_AVG_Drop) and os.access(Spectrum_AVG_Drop, os.R_OK):
	os.remove(Spectrum_AVG_Drop)

if os.path.isfile(Waterfall_AVG_Drop) and os.access(Waterfall_AVG_Drop, os.R_OK):
	os.remove(Waterfall_AVG_Drop)

###########################################################

if not os.path.exists(dir_path+"\Maintenance Ticket Files"):
    os.makedirs(dir_path+"\Maintenance Ticket Files")

###########################################################

PATH = './Tools/Credentials.txt'
if os.path.isfile(PATH) and os.access(PATH, os.R_OK):
    print("The credentials file exists and is readable.")
else:
    print("The credentials file is missing or not readable.")
    time.sleep(3)
    print("Generating new credentials file.")
    credentials = open(dir_path+"/Tools/Credentials.txt", "w+")
    credentials.write("#enter your PEA_username in the line below." + '\n')
    credentials.write('\n')
    credentials.write("#enter your PEA_password in the line below." + '\n')
    credentials.write('\n')
    credentials.write("#enter your Viewpoint_username in the line below." + '\n')
    credentials.write('\n')
    credentials.write("#enter your Viewpoint_password in the line below." + '\n')
    credentials.write('\n')
    credentials.write("#enter your Grafana_username in the line below." + '\n')
    credentials.write('\n')
    credentials.write("#enter your Grafana_password in the line below." + '\n')
    credentials.write('\n')
    credentials.close
    print("Please enter your credentials in the file named 'Credentials.txt'.")
    print('\n')
    # All the stuff inside your window.
    layout = [  [sg.Text("The credentials file is missing or not readable.")],
                [sg.Text("Please enter your credentials in the file named 'Credentials.txt'.")],
                [sg.Text("Press 'Ok' when you are done to continue.")],
                [sg.Button('Ok'), sg.Button('Cancel')]]
    # Create the Window
    window = sg.Window('Missing credentials', layout)
    # Event Loop to process "events" and get the "values" of the inputs
    while True:
	    event, values = window.read()
	    if event == sg.WIN_CLOSED or event == 'Cancel':
		    window.close()
		    sys.exit()
	    window.close()

###########################################################

# Open passwords file
passwords=open(dir_path+"/Tools/Credentials.txt", "r")
# Set password vars
lines=passwords.readlines()
PEA_username=lines[1]
PEA_password=lines[3]
Viewpoint_username=lines[5]
Viewpoint_password=lines[7]
Grafana_username=lines[9]
Grafana_password=lines[11]
# Close passwords file
passwords.close

###########################################################

# Create input to ask for HUB/Node
layout = [  [sg.Text("Type the Mac Domain for the node you would like to investigate.")],
			[sg.Text("For example \"7:0/0\""), sg.InputText()],
            [sg.Button('Ok'), sg.Button('Cancel')] ]

# Create the Window
window = sg.Window('Enter Mac Domain', layout)
# Event Loop to process "events" and get the "values" of the inputs
while True:
    event, value_list = window.read()
    if event == sg.WIN_CLOSED or event == 'Cancel':
	    window.close()
	    sys.exit()
    Mac_Domain = value_list[0]
    break

window.close()

###########################################################

# Convert HUB/Node to MD
# https://kb.ss-cae.net/pages/viewpage.action?pageId=33096083#tab-Alexis
if Mac_Domain == "7:0/0":
	Short_HUB_Input = "ALX 1P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:0/1":
	Short_HUB_Input = "ALX 1P4"
	Cluster = "COS-1"
elif Mac_Domain == "7:0/2":
	Short_HUB_Input = "ALX 2"
	Cluster = "COS-1"
elif Mac_Domain == "7:0/3":
	Short_HUB_Input = "ALX 3P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:0/4":
	Short_HUB_Input = "ALX 3P4"
	Cluster = "COS-1"
elif Mac_Domain == "7:0/5":
	Short_HUB_Input = "ALX 4"
	Cluster = "COS-1"
elif Mac_Domain == "7:0/6":
	Short_HUB_Input = "ALX 5P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:0/7":
	Short_HUB_Input = "ALX 6"
	Cluster = "COS-1"
elif Mac_Domain == "7:0/8":
	Short_HUB_Input = "ALX 7"
	Cluster = "COS-1"
elif Mac_Domain == "7:0/9":
	Short_HUB_Input = "ALX 8P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:0/10":
	Short_HUB_Input = "ALX 8P4"
	Cluster = "COS-1"
elif Mac_Domain == "7:0/11":
	Short_HUB_Input = "ALX 9P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:1/0":
	Short_HUB_Input = "ALX 9P4"
	Cluster = "COS-1"
elif Mac_Domain == "7:1/1":
	Short_HUB_Input = "ALX 10P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:1/2":
	Short_HUB_Input = "ALX 10P4"
	Cluster = "COS-1"
elif Mac_Domain == "7:1/3":
	Short_HUB_Input = "ALX 11P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:1/4":
	Short_HUB_Input = "ALX 11P4"
	Cluster = "COS-1"
elif Mac_Domain == "7:1/5":
	Short_HUB_Input = "ALX 12"
	Cluster = "COS-1"
elif Mac_Domain == "7:1/6":
	Short_HUB_Input = "ALX 13"
	Cluster = "COS-1"
elif Mac_Domain == "7:1/7":
	Short_HUB_Input = "ALX 14"
	Cluster = "COS-1"
elif Mac_Domain == "7:1/8":
	Short_HUB_Input = "ALX 15P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:1/9":
	Short_HUB_Input = "ALX 15P4"
	Cluster = "COS-1"
elif Mac_Domain == "7:1/10":
	Short_HUB_Input = "ALX 16"
	Cluster = "COS-1"
elif Mac_Domain == "7:1/11":
	Short_HUB_Input = "ALX 17P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:2/0":
	Short_HUB_Input = "ALX 17P4"
	Cluster = "COS-1"
elif Mac_Domain == "7:2/1":
	Short_HUB_Input = "ALX 18P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:2/2":
	Short_HUB_Input = "ALX 18P4"
	Cluster = "COS-1"
elif Mac_Domain == "7:2/3":
	Short_HUB_Input = "ALX 19"
	Cluster = "COS-1"
elif Mac_Domain == "7:2/4":
	Short_HUB_Input = "ALX 20"
	Cluster = "COS-1"
elif Mac_Domain == "7:2/5":
	Short_HUB_Input = "ALX 21"
	Cluster = "COS-1"
elif Mac_Domain == "7:2/6":
	Short_HUB_Input = "ALX 22"
	Cluster = "COS-1"
elif Mac_Domain == "7:2/7":
	Short_HUB_Input = "ALX 23P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:2/8":
	Short_HUB_Input = "ALX 23P4"
	Cluster = "COS-1"
elif Mac_Domain == "7:2/9":
	Short_HUB_Input = "ALX 24"
	Cluster = "COS-1"
elif Mac_Domain == "7:2/10":
	Short_HUB_Input = "ALX 25P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:2/11":
	Short_HUB_Input = "ALX 25P4"
	Cluster = "COS-1"
elif Mac_Domain == "7:3/0":
	Short_HUB_Input = "ALX 26P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:3/1":
	Short_HUB_Input = "ALX 26P4"
	Cluster = "COS-1"
elif Mac_Domain == "7:3/2":
	Short_HUB_Input = "ALX 27P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:3/3":
	Short_HUB_Input = "ALX 27P4"
	Cluster = "COS-1"
elif Mac_Domain == "7:3/4":
	Short_HUB_Input = "ALX 28P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:3/5":
	Short_HUB_Input = "ALX 28P4"
	Cluster = "COS-1"
elif Mac_Domain == "7:3/6":
	Short_HUB_Input = "ALX 29"
	Cluster = "COS-1"
elif Mac_Domain == "7:3/7":
	Short_HUB_Input = "ALX 30P1"
	Cluster = "COS-1"
elif Mac_Domain == "7:3/8":
	Short_HUB_Input = "ALX 30P4"
	Cluster = "COS-1"
elif Mac_Domain == "7:3/9":
	Short_HUB_Input = "ALX 31"
	Cluster = "COS-1"
elif Mac_Domain == "7:3/10":
	Short_HUB_Input = "ALX 32"
	Cluster = "COS-1"
elif Mac_Domain == "7:3/11":
	Short_HUB_Input = "ALX 33"
	Cluster = "COS-1"
elif Mac_Domain == "7:4/0":
	Short_HUB_Input = "ALX 34"
	Cluster = "COS-2"
elif Mac_Domain == "7:4/1":
	Short_HUB_Input = "ALX 35P1"
	Cluster = "COS-2"
elif Mac_Domain == "7:4/2":
	Short_HUB_Input = "ALX 35P4"
	Cluster = "COS-2"
elif Mac_Domain == "7:4/3":
	Short_HUB_Input = "ALX 36"
	Cluster = "COS-2"
elif Mac_Domain == "7:4/4":
	Short_HUB_Input = "ALX 37"
	Cluster = "COS-2"
elif Mac_Domain == "7:4/5":
	Short_HUB_Input = "ALX 38"
	Cluster = "COS-2"
elif Mac_Domain == "7:4/6":
	Short_HUB_Input = "ALX 39"
	Cluster = "COS-2"
elif Mac_Domain == "7:4/7":
	Short_HUB_Input = "ALX 40"
	Cluster = "COS-2"
elif Mac_Domain == "7:4/8":
	Short_HUB_Input = "ALX 41"
	Cluster = "COS-2"
elif Mac_Domain == "7:4/9":
	Short_HUB_Input = "ALX 42"
	Cluster = "COS-2"
elif Mac_Domain == "7:4/10":
	Short_HUB_Input = "ALX 43"
	Cluster = "COS-2"
elif Mac_Domain == "7:4/11":
	Short_HUB_Input = "ALX 44"
	Cluster = "COS-2"
elif Mac_Domain == "7:5/0":
	Short_HUB_Input = "ALX 45"
	Cluster = "COS-2"
elif Mac_Domain == "7:5/1":
	Short_HUB_Input = "ALX 46"
	Cluster = "COS-2"
elif Mac_Domain == "7:5/2":
	Short_HUB_Input = "ALX 47"
	Cluster = "COS-2"
elif Mac_Domain == "7:5/3":
	Short_HUB_Input = "ALX 48"
	Cluster = "COS-2"
elif Mac_Domain == "7:5/4":
	Short_HUB_Input = "ALX 49"
	Cluster = "COS-2"
elif Mac_Domain == "7:5/5":
	Short_HUB_Input = "ALX 50P1"
	Cluster = "COS-2"
elif Mac_Domain == "7:5/6":
	Short_HUB_Input = "ALX 50P4"
	Cluster = "COS-2"
elif Mac_Domain == "7:5/7":
	Short_HUB_Input = "ALX 51P1"
	Cluster = "COS-2"
elif Mac_Domain == "7:5/8":
	Short_HUB_Input = "ALX 51P4"
	Cluster = "COS-2"
elif Mac_Domain == "7:5/9":
	Short_HUB_Input = "ALX 52"
	Cluster = "COS-2"
elif Mac_Domain == "7:5/10":
	Short_HUB_Input = "ALX 53"
	Cluster = "COS-2"
elif Mac_Domain == "7:5/11":
	Short_HUB_Input = "ALX 54"
	Cluster = "COS-2"
elif Mac_Domain == "7:9/0":
	Short_HUB_Input = "ALX 55"
	Cluster = "COS-2"
elif Mac_Domain == "7:9/1":
	Short_HUB_Input = "ALX 56P1"
	Cluster = "COS-2"
elif Mac_Domain == "7:9/2":
	Short_HUB_Input = "ALX 56P4"
	Cluster = "COS-2"
elif Mac_Domain == "7:9/3":
	Short_HUB_Input = "ALX 57"
	Cluster = "COS-2"
elif Mac_Domain == "7:9/4":
	Short_HUB_Input = "ALX 58P1"
	Cluster = "COS-2"
elif Mac_Domain == "7:9/5":
	Short_HUB_Input = "ALX 58P4"
	Cluster = "COS-2"
elif Mac_Domain == "7:9/6":
	Short_HUB_Input = "ALX 59"
	Cluster = "COS-2"
elif Mac_Domain == "7:9/7":
	Short_HUB_Input = "ALX 60"
	Cluster = "COS-2"
elif Mac_Domain == "7:9/8":
	Short_HUB_Input = "ALX 61P1"
	Cluster = "COS-2"
elif Mac_Domain == "7:9/9":
	Short_HUB_Input = "ALX 62P1"
	Cluster = "COS-2"
elif Mac_Domain == "7:9/10":
	Short_HUB_Input = "ALX 62P4"
	Cluster = "COS-2"
elif Mac_Domain == "7:9/11":
	Short_HUB_Input = "ALX 63"
	Cluster = "COS-2"
elif Mac_Domain == "7:10/0":
	Short_HUB_Input = "ALX "
	Cluster = "COS-2"
elif Mac_Domain == "7:10/1":
	Short_HUB_Input = "ALX 5P4"
	Cluster = "COS-2"
elif Mac_Domain == "7:10/2":
	Short_HUB_Input = "ALX 61P4"
	Cluster = "COS-2"
elif Mac_Domain == "7:10/3":
	Short_HUB_Input = "ALX 18P3"
	Cluster = "COS-2"
elif Mac_Domain == "7:10/4":
	Short_HUB_Input = "ALX 18P6"
	Cluster = "COS-2"
elif Mac_Domain == "7:10/5":
	Short_HUB_Input = "ALX "
	Cluster = "COS-2"
elif Mac_Domain == "7:10/6":
	Short_HUB_Input = "ALX "
	Cluster = "COS-2"
elif Mac_Domain == "7:10/7":
	Short_HUB_Input = "ALX "
	Cluster = "COS-2"
elif Mac_Domain == "7:10/8":
	Short_HUB_Input = "ALX "
	Cluster = "COS-2"
elif Mac_Domain == "7:10/9":
	Short_HUB_Input = "ALX "
	Cluster = "COS-2"
elif Mac_Domain == "7:10/10":
	Short_HUB_Input = "ALX "
	Cluster = "COS-2"
elif Mac_Domain == "7:10/11":
	Short_HUB_Input = "ALX "
	Cluster = "COS-2"

###########################################################

# Convert input to full HUB name for Viewpoint
Input_Starts_With_ALX = Short_HUB_Input.startswith("ALX")
Input_Starts_With_ARL = Short_HUB_Input.startswith("ARL")
Input_Starts_With_BED = Short_HUB_Input.startswith("BED")
Input_Starts_With_ET = Short_HUB_Input.startswith("ET")
Input_Starts_With_ERI = Short_HUB_Input.startswith("ERI")
Input_Starts_With_ANG = Short_HUB_Input.startswith("ANG")
Input_Starts_With_MAU = Short_HUB_Input.startswith("MAU")
Input_Starts_With_NE = Short_HUB_Input.startswith("NE")
Input_Starts_With_ORG = Short_HUB_Input.startswith("ORG")
Input_Starts_With_OWN = Short_HUB_Input.startswith("OWN")
Input_Starts_With_PER = Short_HUB_Input.startswith("PER")
Input_Starts_With_SPR = Short_HUB_Input.startswith("SPR")
Input_Starts_With_SYL = Short_HUB_Input.startswith("SYL")
Input_Starts_With_UT = Short_HUB_Input.startswith("UT")
Input_Starts_With_WAT = Short_HUB_Input.startswith("WAT")

# Set variable with full name
if Input_Starts_With_ALX == True:
	Viewpoint_Full_HUB = "Alexis"
elif Input_Starts_With_ARL == True:
	Viewpoint_Full_HUB = "Arlington"
elif Input_Starts_With_BED == True:
	Viewpoint_Full_HUB = "Bedford"
elif Input_Starts_With_ET == True:
	Viewpoint_Full_HUB = "East Toledo"
elif Input_Starts_With_ERI == True:
	Viewpoint_Full_HUB = "Erie"
elif Input_Starts_With_ANG == True:
	Viewpoint_Full_HUB = "HE (Angola)"
elif Input_Starts_With_MAU == True:
	Viewpoint_Full_HUB = "Maumee"
elif Input_Starts_With_NE == True:
	Viewpoint_Full_HUB = "NorthEast"
elif Input_Starts_With_ORG == True:
	Viewpoint_Full_HUB = "Oregon"
elif Input_Starts_With_OWN == True:
	Viewpoint_Full_HUB = "Owens"
elif Input_Starts_With_PER == True:
	Viewpoint_Full_HUB = "Perrysburg"
elif Input_Starts_With_SPR == True:
	Viewpoint_Full_HUB = "Springfield"
elif Input_Starts_With_SYL == True:
	Viewpoint_Full_HUB = "Sylvania"
elif Input_Starts_With_UT == True:
	Viewpoint_Full_HUB = "UT"
elif Input_Starts_With_WAT == True:
	Viewpoint_Full_HUB = "Waterville"

# Split Short HUB & Node
HUB_Node_Split = Short_HUB_Input.split()
# Convert to individual lists
Short_HUB_Split_List = [HUB_Node_Split[0]]
Short_Node_Split_List = [HUB_Node_Split[1]]
# Convert to String
out_str = ""
Short_HUB_Split_String = out_str.join(Short_HUB_Split_List)
Short_Node_Split_String = out_str.join(Short_Node_Split_List)

# Add missing 0's to the Node for Viewpoint
if len(Short_Node_Split_String) == 3:
    Corrected_Short_Node_Split = "0"+Short_Node_Split_String
elif len(Short_Node_Split_String) == 1:
    Corrected_Short_Node_Split = "0"+Short_Node_Split_String
else:
	Corrected_Short_Node_Split = Short_Node_Split_String

# Merge corrected short HUB & Node for Viewpoint
Viewpoint_Short_HUB_Node = Short_HUB_Split_String+" "+Corrected_Short_Node_Split

# Add "Node " infront of shortened HUB for PEA.
PEA_Short_HUB_Node = "Node "+Short_HUB_Input

# Convert the "/" in the input to a ":" for easier splitting
Mac_Domain_Corrected = Mac_Domain.replace("/", ":")
# Split the Mac Domain into individual numbers
Mac_Domain_Split = Mac_Domain_Corrected.split(":")
# Place the individual numbers into variables for manipulation
Mac_Domain_Split_First_Number = [Mac_Domain_Split[0]]
Mac_Domain_Split_Second_Number = [Mac_Domain_Split[1]]
Mac_Domain_Split_Third_Number = [Mac_Domain_Split[2]]
# Convert split variables to strings
out_str = ""
Mac_Domain_Split_First_Number_String = out_str.join(Mac_Domain_Split_First_Number)
Mac_Domain_Split_Second_Number_String = out_str.join(Mac_Domain_Split_Second_Number)
Mac_Domain_Split_Third_Number_String = out_str.join(Mac_Domain_Split_Third_Number)

###########################################################

# Open chrome
chromeOptions = webdriver.ChromeOptions()
#chromeOptions.add_argument("--window-size=1600,600")
chromedriver = dir_path+"/Tools/Chrome_Driver/chromedriver.exe"
browser = webdriver.Chrome(executable_path=chromedriver, chrome_options=chromeOptions)
# First Tab
browser.get(("https://harmonicinc.okta.com/login/login.htm?fromURI=/oauth2/v1/authorize/redirect?okta_key=3ZRZV6ACBLrdKnZEm6VZM1vShPqJ5nwPTs7cnMeLObc"))
# Log into Grafana
# Type username
Grafana_username_element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/div/div[2]/div/div/form/div[1]/div[2]/div[1]/div[2]/span/input")))
Grafana_username_element.send_keys(Grafana_username)
# Type password
Grafana_username_element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/div/div[2]/div/div/form/div[1]/div[2]/div[2]/div[2]/span/input")))
Grafana_username_element.send_keys(Grafana_password)
# Click login
# Add extra time for canvas to load
time.sleep(3)
Grafana_signin_element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/div/div[2]/div/div/form/div[2]/input")))
Grafana_signin_element.click()
# Second tab
browser.execute_script("window.open('about:blank', 'tab2');")
browser.switch_to.window("tab2")
browser.get(("https://buckeye.cableos-operations.com/d/core-mac-domain-counters/core-mac-domain-counters?orgId=1&from=now-24h&to=now&var-ds_zabbix=default&var-ds_postgres=ZabbixDirect&var-group=buckeye&var-setup="+Short_HUB_Split_String+"-"+Cluster+"&var-MD=Md"+Mac_Domain_Split_First_Number_String+":"+Mac_Domain_Split_Second_Number_String+"%2F"+Mac_Domain_Split_Third_Number_String+".0"))
# Third tab
browser.execute_script("window.open('about:blank', 'tab3');")
browser.switch_to.window("tab3")
browser.get(("https://buckeye.cableos-operations.com/d/core-cm-states-per-mac-domain/core-cm-states-per-mac-domain?orgId=1&from=now-24h&to=now&var-ds_zabbix=default&var-ds_postgres=ZabbixDirect&var-group=buckeye&var-setup="+Short_HUB_Split_String+"-"+Cluster+"&var-md=Md"+Mac_Domain_Split_First_Number_String+":"+Mac_Domain_Split_Second_Number_String+"%2F"+Mac_Domain_Split_Third_Number_String+".0"))
# Fourth tab
browser.execute_script("window.open('about:blank', 'tab4');")
browser.switch_to.window("tab4")
browser.get(("https://buckeye.cableos-operations.com/d/core-upstream-metrics-mh/core-upstream-metrics?orgId=1&refresh=1m&from=now-24h&to=now&var-ds_zabbix=default&var-ds_postgres=ZabbixDirect&var-group=buckeye&var-setup="+Short_HUB_Split_String+"-"+Cluster+"&var-us_rf_port=Us"+Mac_Domain_Split_First_Number_String+":"+Mac_Domain_Split_Second_Number_String+"%2F"+Mac_Domain_Split_Third_Number_String))
# Fifth tab
browser.execute_script("window.open('about:blank', 'tab5');")
browser.switch_to.window("tab5")
browser.get(("http://10.6.10.12/ViewPoint/site/Site/Login"))
# Log into Viewpoint
# Type username
Viewpoint_username_element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div/div[3]/input[1]")))
Viewpoint_username_element.send_keys(Viewpoint_username)
# Type password
Viewpoint_password_element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div/div[3]/input[2]")))
Viewpoint_password_element.send_keys(Viewpoint_password)
# Click "login"
loginButton = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div/div[3]/div[1]/a")))
loginButton.click()
# Click "RPM"
RPMButton = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[3]/div[1]/div[1]/div/div/div/ul/ul/li/a[1]")))
RPMButton.click()
# Click "Buckeye Cable"
BuckeyeCableButton = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[3]/div[1]/div[1]/div/div/div/ul/ul/ul/li/a[1]")))
BuckeyeCableButton.click()
# Click "Alexis"
Full_HUB = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[text()="%s"]' % Viewpoint_Full_HUB)))
Full_HUB.click()
# Click "Node"
Shortened_Node = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[text()="%s"]' % Viewpoint_Short_HUB_Node)))
Shortened_Node.click()
# Click "Return Spectrum"
Return_Spectrum_Button = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[4]/div[1]/div/div/div/div/ul/li")))
Return_Spectrum_Button.click()	
# Click "Mode"
Mode_Dropdown = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[2]/li[2]/select")))
Mode_Dropdown.click()
# Click "Historical"
Historical_Dropdown = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[2]/li[2]/select/option[2]")))
Historical_Dropdown.click()
# Click "Display"
Display_Dropdown = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[3]/ul[1]/li[2]/select")))
Display_Dropdown.click()
# Click "Spectrum"
Spectrum_Dropdown_Option = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[3]/ul[1]/li[2]/select/option[1]")))
Spectrum_Dropdown_Option.click()
# Close First tab
browser.switch_to.window(browser.window_handles[0])
browser.close()

###########################################################

# Create input to ask for HUB/Node
layout = [  [sg.Text("Please review the open tabs to determine if a maintenance ticket is required.")],
            [sg.Text("Then, press 'Ok' to scrape the node data if one is required.")],
            [sg.Button('Ok'), sg.Button('Exit')] ]

# Create the Window
window = sg.Window('Maintenance ticket?', layout)
# Event Loop to process "events" and get the "values" of the inputs
while True:
    event, value_list = window.read()
    if event == sg.WIN_CLOSED or event == 'Exit':
	    window.close()
	    quit_chrome()
	    sys.exit()
    break

window.close()

###########################################################

# Open chrome
chromeOptions = webdriver.ChromeOptions()
prefs = {"download.default_directory" : dir_path+"\Maintenance Ticket Files\\"}
chromeOptions.add_experimental_option("prefs",prefs)
#chromeOptions.add_argument("--window-size=1600,600")
chromedriver = dir_path+"/Tools/Chrome_Driver/chromedriver.exe"
browser = webdriver.Chrome(executable_path=chromedriver, chrome_options=chromeOptions)
# Open PEA in Chrome.
browser.get(("https://buckeye.pea.zcorum.com/pnm4/index.php/detailView"))

# Grab screenshot of the PEA map for the node & export the bad visible modems.
try:
	# Log into PEA	
	# Type username
	PEA_username_element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div/div/div/div/div[5]/form/div[1]/div/input")))
	PEA_username_element.send_keys(PEA_username)
	
	# Type password
	PEA_username_element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div/div/div/div/div[5]/form/div[2]/div/input")))
	PEA_username_element.send_keys(PEA_password)
	
	# Enter HUB in the "HUB" field
	# Add extra time for page to load
	time.sleep(3)
	HUB_Field = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[4]/div[6]/div[3]/div/div[1]/div/div/div[2]/div[2]/div/input[2]")))
	HUB_Field.send_keys(PEA_Short_HUB_Node)
	
	# Click the suggested Node
	# Add extra time for page to load
	time.sleep(1)
	suggestButton = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[4]/div[6]/div[3]/div/div[1]/div/div/div[2]/div[2]/div/select/option[1]")))
	suggestButton.click()
	
	# Click the search button
	# Add extra time for page to load
	time.sleep(1)
	searchButton = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[4]/div[6]/div[3]/div/div[1]/div/div/div[2]/div[2]/button")))
	searchButton.click()
	
	# Click the upstream search button
	upstreamsearchButton = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[7]/div[2]/div/div[3]/button[2]")))
	upstreamsearchButton.click()
	
	# Click the red button
	redButton = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[4]/div[6]/div[3]/div/div[1]/div/div/div[2]/div[4]/span/input[1]")))
	redButton.click()

	# Take screenshot of the map
	# Add extra time for canvas to load
	time.sleep(3)
	element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[4]/div[6]/div[3]/div/div[3]/div/div[1]/div[2]/div/div/div/div[1]/div[1]/canvas")))
	location = element.location
	size = element.size
	browser.save_screenshot(dir_path+"/Maintenance Ticket Files/Full_Map_Element.png")
	# Crop screenshot of the map
	x = location['x']
	y = location['y']
	width = location['x']+size['width']
	height = location['y']+size['height']
	im = Image.open(dir_path+"/Maintenance Ticket Files/Full_Map_Element.png")
	im = im.crop((int(x), int(y), int(width), int(height)))
	im.save(dir_path+"/Maintenance Ticket Files/PEA critical modem map.png")
	
	# Click on data tab
	dataTab = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[4]/div[6]/div[3]/div/div[3]/ul/li[2]/a")))
	dataTab.click()
	
	# Click on sorting
	VTDR_Element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "//th[contains(text(),'Outside Plant')]")))
	VTDR_Element.click()
	time.sleep(2)
	VTDR_Element.click()
	
	# Wait for page to load
	time.sleep(4)
	
	# Click on Export
	Export_Element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[4]/div[6]/div[3]/div/div[3]/div/div[2]/div/div/div[1]/div[6]/div/button")))
	Export_Element.click()

	# Click on Only Visible Data
	Only_Visible_Data_Element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[4]/div[6]/div[3]/div/div[3]/div/div[2]/div/div/div[1]/div[6]/div/ul/li[2]/a")))
	Only_Visible_Data_Element.click()
	
	# Click the yellow button
	yellowButton = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[4]/div[6]/div[3]/div/div[1]/div/div/div[2]/div[4]/span/input[2]")))
	yellowButton.click()
	
	# Click on data tab
	dataTab = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[4]/div[6]/div[3]/div/div[3]/ul/li[2]/a")))
	dataTab.click()
	
	# Click on sorting
	VTDR_Element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "//th[contains(text(),'Outside Plant')]")))
	VTDR_Element.click()
	time.sleep(2)
	VTDR_Element.click()
	
	# Wait for page to load
	time.sleep(4)
	
	# Click on Export
	Export_Element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[4]/div[6]/div[3]/div/div[3]/div/div[2]/div/div/div[1]/div[6]/div/button")))
	Export_Element.click()

	# Click on Only Visible Data
	Only_Visible_Data_Element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[4]/div[6]/div[3]/div/div[3]/div/div[2]/div/div/div[1]/div[6]/div/ul/li[2]/a")))
	Only_Visible_Data_Element.click()
	
	#### Quit Browser
	time.sleep(2)
	quit_chrome()
except Exception:
	quit_chrome_exception()
	sys.exit()

# Convert CSV to XLSX
wb = Workbook()
ws = wb.active
with open(dir_path+"\Maintenance Ticket Files\ModemReport.csv", 'r') as f:
    for row in csv.reader(f):
        ws.append(row)
wb.save(dir_path+"\Maintenance Ticket Files\ModemReport_Red.xlsx")

with open(dir_path+"\Maintenance Ticket Files\ModemReport (1).csv", 'r') as f:
    for row in csv.reader(f):
        ws.append(row)
wb.save(dir_path+"\Maintenance Ticket Files\ModemReport_RedYellow.xlsx")

# Remove unecessary columns for PEA vTDR
wb = openpyxl.load_workbook(dir_path+"\Maintenance Ticket Files\ModemReport_RedYellow.xlsx")
ws = wb.active
# Delete B,C,D
ws.delete_cols(2,3)
# Delete O-U
ws.delete_cols(12,7)
# Delete W-AG
ws.delete_cols(13,11)
# Save XLSX file
wb.save(dir_path+"\Maintenance Ticket Files\PEA vTDR.xlsx")

# Remove unecessary columns for critical modems
wb = openpyxl.load_workbook(dir_path+"\Maintenance Ticket Files\ModemReport_Red.xlsx")
ws = wb.active
# Delete B-U
ws.delete_cols(2,20)
# Delete C-M
ws.delete_cols(3,11)
# Save XLSX file
wb.save(dir_path+"\Maintenance Ticket Files\PEA critical modems.xlsx")

# Delete extra files
os.remove(dir_path+"\Maintenance Ticket Files\ModemReport.csv")
os.remove(dir_path+"\Maintenance Ticket Files\ModemReport (1).csv")
os.remove(dir_path+"\Maintenance Ticket Files\ModemReport_Red.xlsx")
os.remove(dir_path+"\Maintenance Ticket Files\ModemReport_RedYellow.xlsx")
os.remove(dir_path+"\Maintenance Ticket Files\Full_Map_Element.png")

###########################################################

# Open chrome
chromeOptions = webdriver.ChromeOptions()
prefs = {"download.default_directory" : dir_path+"\Maintenance Ticket Files\\"}
chromeOptions.add_experimental_option("prefs",prefs)
#chromeOptions.add_argument("--window-size=1600,600")
chromedriver = dir_path+"/Tools/Chrome_Driver/chromedriver.exe"
browser = webdriver.Chrome(executable_path=chromedriver, chrome_options=chromeOptions)
# Open Viewpoint in Chrome.
browser.get(("http://10.6.10.12/ViewPoint/site/Site/Login"))

# 
try:
	# Log into Viewpoint
	# Type username
	Viewpoint_username_element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div/div[3]/input[1]")))
	Viewpoint_username_element.send_keys(Viewpoint_username)

	# Type password
	Viewpoint_password_element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div/div[3]/input[2]")))
	Viewpoint_password_element.send_keys(Viewpoint_password)

	# Click "login"
	loginButton = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div/div[3]/div[1]/a")))
	loginButton.click()

	# Click "RPM"
	RPMButton = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[3]/div[1]/div[1]/div/div/div/ul/ul/li/a[1]")))
	RPMButton.click()

	# Click "Buckeye Cable"
	BuckeyeCableButton = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[3]/div[1]/div[1]/div/div/div/ul/ul/ul/li/a[1]")))
	BuckeyeCableButton.click()

	# Click "Alexis"
	Full_HUB = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[text()="%s"]' % Viewpoint_Full_HUB)))
	Full_HUB.click()

	# Click "Node"
	Shortened_Node = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[text()="%s"]' % Viewpoint_Short_HUB_Node)))
	Shortened_Node.click()

	# Click "Return Spectrum"
	Return_Spectrum_Button = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[4]/div[1]/div/div/div/div/ul/li")))
	Return_Spectrum_Button.click()
	
	# Click "Mode"
	Mode_Dropdown = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[2]/li[2]/select")))
	Mode_Dropdown.click()

	# Click "Historical"
	Historical_Dropdown = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[2]/li[2]/select/option[2]")))
	Historical_Dropdown.click()

	# Click "Display"
	Display_Dropdown = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[3]/ul[1]/li[2]/select")))
	Display_Dropdown.click()

	# Click "Spectrum"
	Spectrum_Dropdown_Option = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[3]/ul[1]/li[2]/select/option[1]")))
	Spectrum_Dropdown_Option.click()

	# Click "Back button for 15 minute increments"
	Back_15_Minute_Button = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[4]/div[1]/div/div/div/div[2]/ul/li[2]")))
	Back_15_Minute_Button.click()

	# Take screenshot of the average spectrum
	# Add extra time for canvas to load
	time.sleep(3)
	element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[4]/div[1]/div/div/div/div[4]/div/div/div/div[1]")))
	location = element.location
	size = element.size
	browser.save_screenshot(Full_Spectrum_AVG_Drop)
	# Crop screenshot of the average spectrum
	x = location['x']
	y = location['y']
	width = location['x']+size['width']
	height = location['y']+size['height']
	im = Image.open(Full_Spectrum_AVG_Drop)
	im = im.crop((int(x), int(y), int(width), int(height)))
	im.save(Spectrum_AVG_Drop)

	# Click "Active Trace"
	Active_Trace_Dropdown = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[3]/ul[2]/ul/li[2]/select")))
	Active_Trace_Dropdown.click()

	# Click "MAX"
	Max_Dropdown_Option = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[3]/ul[2]/ul/li[2]/select/option[1]")))
	Max_Dropdown_Option.click()

	# Take screenshot of the max spectrum
	# Add extra time for canvas to load
	time.sleep(3)
	element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[4]/div[1]/div/div/div/div[4]/div/div/div/div[1]")))
	location = element.location
	size = element.size
	browser.save_screenshot(Full_Spectrum_MAX_Drop)
	# Crop screenshot of the max spectrum
	x = location['x']
	y = location['y']
	width = location['x']+size['width']
	height = location['y']+size['height']
	im = Image.open(Full_Spectrum_MAX_Drop)
	im = im.crop((int(x), int(y), int(width), int(height)))
	im.save(Spectrum_MAX_Drop)

	# Click "Display"
	Display_Dropdown = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[3]/ul[1]/li[2]/select")))
	Display_Dropdown.click()

	# Click "Waterfall"
	Waterfall_Dropdown_Option = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[3]/ul[1]/li[2]/select/option[2]")))
	Waterfall_Dropdown_Option.click()

	# Click "Time Span"
	Time_Span_Dropdown = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[3]/ul[1]/li[4]/select")))
	Time_Span_Dropdown.click()

	# Click "24 Hr"
	Twenty_four_hours_dropdown_option = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[3]/ul[1]/li[4]/select/option[5]")))
	Twenty_four_hours_dropdown_option.click()

	# Take screenshot of the 24 hr MAX waterfall
	# Add extra time for canvas to load
	time.sleep(3)
	element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[4]/div[1]/div/div/div/div[4]/canvas[1]")))
	location = element.location
	size = element.size
	browser.save_screenshot(Full_Waterfall_MAX_Drop)
	# Crop screenshot of the 24 hr MAX waterfall
	x = location['x']
	y = location['y']
	width = location['x']+size['width']
	height = location['y']+size['height']
	im = Image.open(Full_Waterfall_MAX_Drop)
	im = im.crop((int(x), int(y), int(width), int(height)))
	im.save(Waterfall_MAX_Drop)

	# Click "Active Trace"
	Active_Trace_Dropdown = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[3]/ul[2]/ul/li[2]/select")))
	Active_Trace_Dropdown.click()

	# Click "AVG"
	Average_Dropdown_Option = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[2]/div[1]/div/div[2]/div/ul[3]/ul[2]/ul/li[2]/select/option[2]")))
	Average_Dropdown_Option.click()
	
	# Take screenshot of the 24 hr AVG waterfall
	# Add extra time for canvas to load
	time.sleep(3)
	element = WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/div[2]/div[4]/div[1]/div/div/div/div[4]/canvas[1]")))
	location = element.location
	size = element.size
	browser.save_screenshot(Full_Waterfall_AVG_Drop)
	# Crop screenshot of the 24 hr AVG waterfall
	x = location['x']
	y = location['y']
	width = location['x']+size['width']
	height = location['y']+size['height']
	im = Image.open(Full_Waterfall_AVG_Drop)
	im = im.crop((int(x), int(y), int(width), int(height)))
	im.save(Waterfall_AVG_Drop)

	# Delete extra files
	os.remove(Full_Spectrum_MAX_Drop)
	os.remove(Full_Spectrum_AVG_Drop)
	os.remove(Full_Waterfall_MAX_Drop)
	os.remove(Full_Waterfall_AVG_Drop)

	#### Quit Browser
	quit_chrome()
except Exception:
	quit_chrome_exception()
	sys.exit()